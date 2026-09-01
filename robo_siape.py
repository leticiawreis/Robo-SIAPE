import csv
import logging
import os
import shutil
import sys
import tempfile
import time
import zipfile

from datetime import datetime

import requests
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _obter_pasta_base():
    """Retorna a pasta raiz do projeto (Robo-SIAPE/), onde 'saida/' deve nascer.

    Rodando como script (.py), essa é a pasta onde interface.py/robo_siape.py
    estão de fato.

    Rodando como .exe (PyInstaller --onefile), o executável sempre fica
    dentro de dist/ (ex.: Robo-SIAPE/dist/RoboSIAPE.exe) — então subimos
    mais um nível a partir do .exe para chegar na raiz do projeto
    (Robo-SIAPE/), em vez de deixar a saida/ nascer dentro de dist/.

    Retorna:
        str: caminho absoluto da pasta raiz do projeto.
    """
    if getattr(sys, "frozen", False):
        pasta_dist = os.path.dirname(os.path.abspath(sys.executable))
        return os.path.dirname(pasta_dist)
    return os.path.dirname(os.path.abspath(__file__))


PASTA_SAIDA = os.path.join(_obter_pasta_base(), "saida")
PASTA_EXTRAIDA = os.path.join(PASTA_SAIDA, "_processamento")
MAX_TENTATIVAS_DOWNLOAD = 3

# Índices mantidos por compatibilidade com telas antigas (não são mais usados
# para navegação em navegador — o download agora é direto via requests).
INDICES_ANOS = {
    "2026": 1, "2025": 2, "2024": 3, "2023": 4, "2022": 5,
    "2021": 6, "2020": 7, "2019": 8, "2018": 9, "2017": 10,
    "2016": 11, "2015": 12, "2014": 13, "2013": 14,
}

MESES_OPCOES = {
    "janeiro": 1, "fevereiro": 2, "março": 3, "abril": 4,
    "maio": 5, "junho": 6, "julho": 7, "agosto": 8,
    "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12,
}

MESES_2026 = {
    "janeiro", "fevereiro", "março", "abril", "maio", "junho"
}

PALAVRAS_MONETARIAS = [
    "remuneracao", "remuneração", "remunerações", "remuneracoes", "abate-teto",
    "gratificação", "gratificacao", "férias", "ferias", "irrf", "pss", "rpgs",
    "dedução", "deducao", "deduções", "deducoes", "pensão", "pensao", "fundo",
    "taxa", "verbas", "total"

]

FORMATO_MOEDA = "R$ #,##0.00"

CABECALHOS_HTTP = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36"
    )
}


class QueueLogHandler(logging.Handler):
    """Handler customizado do módulo `logging` que, em vez de escrever em
    arquivo ou console, envia cada mensagem de log formatada para uma
    fila (queue.Queue) fornecida na criação.

    É assim que os logs do pipeline (que roda em uma thread de trabalho)
    chegam em tempo real até a interface gráfica, que consome essa fila
    e exibe as mensagens no console de log da tela de execução.
    """

    def __init__(self, log_queue):
        """Guarda a referência da fila onde as mensagens de log serão
        colocadas.

        Parâmetros:
            log_queue (queue.Queue): fila compartilhada com o
                consumidor (ex: a interface gráfica).
        """
        super().__init__()
        self.log_queue = log_queue

    def emit(self, record):
        """Formata o registro de log recebido e o coloca na fila, como
        uma tupla ("log", mensagem_formatada).

        Chamado automaticamente pelo framework `logging` sempre que uma
        mensagem é registrada através de um logger que usa este handler.
        Se ocorrer algum erro ao formatar/enfileirar, delega o
        tratamento ao mecanismo padrão de erro do `logging`
        (self.handleError).

        Parâmetros:
            record (logging.LogRecord): registro de log a ser emitido.
        """
        try:
            mensagem = self.format(record)
            self.log_queue.put(("log", mensagem))
        except Exception:
            self.handleError(record)


def meses_disponiveis(ano):
    """Retorna a lista de meses que podem ser selecionados para um
    determinado ano.

    Para o ano corrente (2026), retorna apenas os meses já decorridos/
    disponíveis (definidos em MESES_2026), já que meses futuros não
    teriam dados publicados. Para os demais anos, retorna todos os 12
    meses do ano.

    Parâmetros:
        ano (str): ano a consultar (ex: "2026").

    Retorna:
        list[str]: lista de nomes de meses (por extenso, em minúsculas)
        disponíveis para aquele ano.
    """
    if ano == "2026":
        return [mes for mes in MESES_OPCOES if mes in MESES_2026]

    return list(MESES_OPCOES.keys())


def criar_pasta_execucao(ano, mes):
    """Cria uma pasta exclusiva para esta execução, sempre nova.

    Mesmo repetindo o mesmo ano/mês, a pasta nunca é reaproveitada nem
    sobrescrita — o nome carrega um carimbo de data/hora e, no
    improvável caso de colisão no mesmo segundo, ganha um sufixo extra.

    Parâmetros:
        ano (str): ano da execução.
        mes (str): mês da execução, por extenso (usado para obter o
            número do mês via MESES_OPCOES).

    Retorna:
        str: caminho absoluto da pasta recém-criada para esta execução,
        dentro de PASTA_SAIDA.
    """
    marca_tempo = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_base = f"{ano}-{MESES_OPCOES[mes]:02d}_{marca_tempo}"

    caminho = os.path.join(PASTA_SAIDA, nome_base)
    sufixo = 1
    while os.path.exists(caminho):
        sufixo += 1
        caminho = os.path.join(PASTA_SAIDA, f"{nome_base}_{sufixo}")

    os.makedirs(caminho, exist_ok=True)
    return caminho


def configurar_logger(ano, mes, log_queue, pasta_execucao):
    """Cria e configura um logger dedicado a uma única execução do
    pipeline, com dois destinos simultâneos para cada mensagem:
    1. Um arquivo de log (.log) dentro da pasta desta execução.
    2. A fila fornecida (log_queue), via QueueLogHandler, para que a
       interface gráfica exiba as mensagens em tempo real.

    O logger recebe um nome único baseado em `id(log_queue)` para evitar
    conflito entre execuções concorrentes/sucessivas, e quaisquer
    handlers pré-existentes nesse logger são removidos antes de
    adicionar os novos (para não duplicar mensagens em reexecuções).

    Parâmetros:
        ano (str): ano da execução (usado no nome do arquivo de log).
        mes (str): mês da execução, por extenso.
        log_queue (queue.Queue): fila para onde as mensagens também
            serão enviadas em tempo real.
        pasta_execucao (str): pasta desta execução, onde o arquivo de
            log será salvo.

    Retorna:
        tuple: (logger, caminho_log) — o logger configurado, pronto
        para uso, e o caminho completo do arquivo de log criado.
    """
    caminho_log = os.path.join(
        pasta_execucao,
        f"execucao_{ano}_{MESES_OPCOES[mes]:02d}.log"
    )

    logger = logging.getLogger(f"robo_siape_{id(log_queue)}")
    logger.setLevel(logging.INFO)
    logger.propagate = False

    for handler in list(logger.handlers):
        logger.removeHandler(handler)
        handler.close()

    arquivo_handler = logging.FileHandler(
        caminho_log,
        encoding="utf-8"
    )

    fila_handler = QueueLogHandler(log_queue)

    formato = logging.Formatter(
        "%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%d/%m/%Y %H:%M:%S"
    )

    arquivo_handler.setFormatter(formato)
    fila_handler.setFormatter(formato)

    logger.addHandler(arquivo_handler)
    logger.addHandler(fila_handler)

    return logger, caminho_log


def criar_pasta_trabalho():
    """Cria uma pasta temporária exclusiva para o download do pacote
    ZIP desta execução (usando o diretório temporário do sistema
    operacional).

    Retorna:
        str: caminho absoluto da pasta temporária criada, com o prefixo
        "robo_siape_".
    """
    return tempfile.mkdtemp(
        prefix="robo_siape_"
    )


def limpar_pasta_trabalho(pasta):
    """Remove (recursivamente) a pasta de trabalho temporária, se ela
    existir, ignorando erros durante a remoção (ex: arquivos em uso).

    Parâmetros:
        pasta (str): caminho da pasta a remover. Se for None ou não
            existir, a função não faz nada.
    """
    if pasta and os.path.exists(pasta):
        shutil.rmtree(pasta, ignore_errors=True)


# =====================================================================
# DOWNLOAD — direto via requests (headless, sem navegador e sem CAPTCHA)
# =====================================================================
def baixar_pacote(ano, mes, tipo, pasta_destino, logger):
    """Baixa o pacote .zip do Portal da Transparência diretamente por HTTP.

    Tenta primeiro o link "oficial" da página de download e, se ele falhar,
    tenta o endereço estático da CGU (destino do redirecionamento), que
    costuma escapar de bloqueios de WAF.

    Durante o download, grava o conteúdo em um arquivo parcial
    (extensão ".parcial") em blocos de 512 KB, registrando no log o
    progresso a cada 10% (quando o tamanho total é informado pelo
    servidor). Ao final, valida se o conteúdo baixado é de fato um ZIP
    válido antes de renomear o arquivo parcial para o nome final.

    Parâmetros:
        ano (str): ano da competência a baixar.
        mes (str): mês da competência, por extenso.
        tipo (str): tipo do pacote de dados (ex: "Servidores_SIAPE").
        pasta_destino (str): pasta onde o arquivo ZIP baixado será salvo.
        logger (logging.Logger): logger usado para registrar o progresso
            e eventuais falhas do download.

    Retorna:
        str: caminho completo do arquivo ZIP baixado com sucesso.

    Levanta:
        FileNotFoundError: se o servidor responder 404 (pacote não
            publicado para essa competência) — nesse caso não adianta
            tentar o segundo endereço.
        RuntimeError: se todas as tentativas de endereço falharem por
            outros motivos (erro de rede, conteúdo inválido, etc.).
    """
    competencia = f"{ano}{MESES_OPCOES[mes]:02d}"
    nome_do_pacote = f"{competencia}_{tipo}"

    enderecos = [
        f"https://portaldatransparencia.gov.br/download-de-dados/servidores/{nome_do_pacote}",
        f"https://dadosabertos-download.cgu.gov.br/PortalDaTransparencia/saida/servidores/{nome_do_pacote}.zip",
    ]

    os.makedirs(pasta_destino, exist_ok=True)
    caminho_final = os.path.join(pasta_destino, f"{nome_do_pacote}.zip")
    caminho_parcial = caminho_final + ".parcial"

    ultimo_erro = None

    for endereco in enderecos:
        logger.info("Baixando de: %s", endereco)

        try:
            with requests.get(
                endereco,
                headers=CABECALHOS_HTTP,
                stream=True,
                timeout=(15, 120),
            ) as resposta:
                if resposta.status_code == 404:
                    raise FileNotFoundError(
                        f"Pacote {nome_do_pacote} não publicado (404)."
                    )
                resposta.raise_for_status()

                tamanho_informado = int(resposta.headers.get("content-length") or 0)
                bytes_gravados = 0
                proximo_marco = 10

                with open(caminho_parcial, "wb") as arquivo:
                    for bloco in resposta.iter_content(chunk_size=512 * 1024):
                        arquivo.write(bloco)
                        bytes_gravados += len(bloco)

                        if tamanho_informado:
                            percentual = (bytes_gravados / tamanho_informado) * 100
                            if percentual >= proximo_marco:
                                logger.info("Progresso do download: %d%%", int(percentual))
                                proximo_marco += 10

                if not zipfile.is_zipfile(caminho_parcial):
                    os.remove(caminho_parcial)
                    raise ValueError("O conteúdo recebido não é um ZIP válido.")

                shutil.move(caminho_parcial, caminho_final)
                logger.info(
                    "Download concluído: %s (%.1f MB)",
                    os.path.basename(caminho_final),
                    bytes_gravados / 1024 / 1024,
                )
                return caminho_final

        except FileNotFoundError:
            # Pacote realmente não existe para essa competência/tipo — não
            # adianta tentar o outro endereço.
            raise
        except (requests.RequestException, ValueError) as erro:
            logger.warning("Falha ao baixar de %s: %s", endereco, erro)
            ultimo_erro = erro
            if os.path.exists(caminho_parcial):
                os.remove(caminho_parcial)

    raise RuntimeError(
        f"Não foi possível baixar o pacote em nenhum endereço. Último erro: {ultimo_erro}"
    )


def baixar_com_retentativas(ano, mes, tipo, logger):
    """Executa o download do pacote (baixar_pacote) com retentativas
    automáticas em caso de falha, até MAX_TENTATIVAS_DOWNLOAD vezes,
    aguardando 2 segundos entre cada tentativa.

    Cria uma pasta de trabalho temporária antes de tentar o download.
    Se o erro for FileNotFoundError (pacote não existe para aquela
    competência), interrompe imediatamente sem novas tentativas — não
    faz sentido tentar de novo algo que não foi publicado. Se todas as
    tentativas se esgotarem por outros motivos, limpa a pasta de
    trabalho e relança o erro como RuntimeError.

    Parâmetros:
        ano (str): ano da competência a baixar.
        mes (str): mês da competência, por extenso.
        tipo (str): tipo do pacote de dados.
        logger (logging.Logger): logger para registrar cada tentativa.

    Retorna:
        tuple: (arquivo, pasta_download) — caminho do ZIP baixado com
        sucesso e caminho da pasta de trabalho temporária usada (que
        deve ser limpa posteriormente pelo chamador).

    Levanta:
        FileNotFoundError: se o pacote não estiver publicado para a
            competência solicitada.
        RuntimeError: se o download falhar em todas as tentativas por
            outros motivos.
    """
    pasta_download = criar_pasta_trabalho()

    for tentativa in range(1, MAX_TENTATIVAS_DOWNLOAD + 1):
        try:
            logger.info(
                "Tentativa de download %d/%d.",
                tentativa,
                MAX_TENTATIVAS_DOWNLOAD
            )

            arquivo = baixar_pacote(ano, mes, tipo, pasta_download, logger)

            logger.info(
                "Download validado na tentativa %d/%d.",
                tentativa,
                MAX_TENTATIVAS_DOWNLOAD
            )

            return arquivo, pasta_download

        except FileNotFoundError:
            limpar_pasta_trabalho(pasta_download)
            raise

        except Exception as erro:
            logger.warning(
                "Falha na tentativa %d/%d: %s",
                tentativa,
                MAX_TENTATIVAS_DOWNLOAD,
                erro
            )

            if tentativa == MAX_TENTATIVAS_DOWNLOAD:
                limpar_pasta_trabalho(pasta_download)
                raise RuntimeError(
                    "O download falhou após 3 tentativas."
                ) from erro

            logger.info("Tentando o download novamente.")
            time.sleep(2)

    raise RuntimeError("Não foi possível realizar o download.")


# =====================================================================
# TRATAMENTO DO CSV / GERAÇÃO DO EXCEL
# =====================================================================
def encontrar_arquivo_remuneracao(pasta_extraida):
    """Procura, dentro da pasta onde o ZIP foi extraído, o arquivo CSV
    que contém os dados de remuneração dos servidores.

    Percorre recursivamente a pasta procurando arquivos cujo nome
    (ignorando maiúsculas/minúsculas) contenha "remuneracao" ou
    "remuneração" e termine em ".csv".

    Parâmetros:
        pasta_extraida (str): pasta onde o conteúdo do ZIP foi
            extraído.

    Retorna:
        str: caminho completo do primeiro CSV de remuneração encontrado.

    Levanta:
        FileNotFoundError: se nenhum arquivo correspondente for
            encontrado dentro da pasta.
    """
    candidatos = []

    for raiz, _, arquivos in os.walk(pasta_extraida):
        for nome in arquivos:
            nome_normalizado = nome.lower()

            if "remuneracao" in nome_normalizado or "remuneração" in nome_normalizado:
                if nome_normalizado.endswith(".csv"):
                    candidatos.append(os.path.join(raiz, nome))

    if not candidatos:
        raise FileNotFoundError(
            "Não encontrei o CSV de remuneração dentro do ZIP."
        )

    return candidatos[0]


def preparar_csv(caminho):
    """Abre o arquivo CSV testando várias codificações (encodings) até
    encontrar uma que consiga ler o conteúdo sem erro, já que os
    arquivos do Portal da Transparência podem vir em codificações
    diferentes (UTF-8 com/sem BOM, CP1252, Latin-1).

    Tenta, nesta ordem: "utf-8-sig", "utf-8", "cp1252", "latin1". Para
    cada uma, tenta ler um trecho inicial do arquivo (4096 bytes) como
    teste; se não houver erro de decodificação, retorna o arquivo já
    reposicionado no início, junto com um leitor CSV configurado com
    delimitador ";".

    Parâmetros:
        caminho (str): caminho do arquivo CSV a abrir.

    Retorna:
        tuple: (arquivo, leitor) — o objeto de arquivo aberto (que deve
        ser fechado pelo chamador) e o csv.reader pronto para iterar
        sobre as linhas.

    Levanta:
        UnicodeDecodeError: se nenhuma das codificações testadas
            conseguir ler o arquivo.
    """
    encodings = ("utf-8-sig", "utf-8", "cp1252", "latin1")

    ultimo_erro = None

    for encoding in encodings:
        arquivo = None
        try:
            arquivo = open(
                caminho,
                "r",
                encoding=encoding,
                newline=""
            )

            arquivo.read(4096)
            arquivo.seek(0)

            leitor = csv.reader(
                arquivo,
                delimiter=";"
            )

            return arquivo, leitor

        except UnicodeDecodeError as erro:
            ultimo_erro = erro
            if arquivo is not None:
                arquivo.close()

    raise UnicodeDecodeError(
        "csv",
        b"",
        0,
        1,
        f"Não foi possível identificar a codificação do CSV: {ultimo_erro}"
    )


def coluna_vazia(valor):
    """Verifica se um valor de célula deve ser considerado "vazio" para
    fins de tratamento (None ou string composta só de espaços).

    Parâmetros:
        valor: valor da célula a verificar (geralmente str ou None).

    Retorna:
        bool: True se o valor for considerado vazio, False caso
        contrário.
    """
    return valor is None or str(valor).strip() == ""


def identificar_colunas_validas(caminho_csv, logger):
    """Faz uma primeira passada pelo CSV para identificar quais colunas
    realmente possuem algum dado preenchido em pelo menos uma linha, e
    quantas linhas não estão totalmente vazias.

    Colunas que estão vazias em todas as linhas do arquivo são
    descartadas do resultado — elas não farão parte da planilha final,
    evitando colunas inúteis no Excel gerado. Linhas totalmente vazias
    (todas as colunas vazias) não são contabilizadas em `total_linhas`.

    Parâmetros:
        caminho_csv (str): caminho do CSV de remuneração a analisar.
        logger (logging.Logger): logger usado para registrar quantas
            colunas válidas e quantos registros foram encontrados.

    Retorna:
        tuple: (cabecalho_limpo, colunas_validas) — cabecalho_limpo é a
        lista dos nomes das colunas que possuem dados (na ordem
        original); colunas_validas é a lista dos índices originais
        dessas colunas no CSV, usada posteriormente para filtrar cada
        linha de dados.

    Levanta:
        ValueError: se o CSV estiver vazio (sem sequer um cabeçalho).
    """
    arquivo, leitor = preparar_csv(caminho_csv)

    try:
        cabecalho = next(leitor, None)

        if not cabecalho:
            raise ValueError(
                "O CSV de remuneração está vazio."
            )

        cabecalho = list(cabecalho)
        preenchidas = [False] * len(cabecalho)
        total_linhas = 0

        for linha in leitor:
            if not any(
                not coluna_vazia(valor)
                for valor in linha
            ):
                continue

            total_linhas += 1

            for indice in range(len(cabecalho)):
                if indice < len(linha) and not coluna_vazia(linha[indice]):
                    preenchidas[indice] = True

        colunas_validas = [
            indice
            for indice, preenchida in enumerate(preenchidas)
            if preenchida
        ]

        cabecalho_limpo = [
            cabecalho[indice]
            for indice in colunas_validas
        ]

        logger.info(
            "Tratamento identificado: %d colunas válidas e %d registros não vazios.",
            len(cabecalho_limpo),
            total_linhas
        )

        return cabecalho_limpo, colunas_validas

    finally:
        arquivo.close()


def converter_valor_monetario(valor):
    """Converte um texto de valor monetário (no formato brasileiro,
    ex: "1.234,56" ou "R$ 1.234,56") para um número float, para que
    possa ser gravado na planilha como número (e não como texto) e
    formatado como moeda.

    Trata dois formatos possíveis:
    - Formato brasileiro com vírgula decimal (ex: "1.234,56"): remove
      os pontos de milhar e troca a vírgula por ponto decimal.
    - Formato sem vírgula, mas com múltiplos pontos (ex: "1.234.567"):
      remove todos os pontos, tratando-os como separadores de milhar.

    Parâmetros:
        valor: texto (ou None) a converter.

    Retorna:
        float | None: o valor numérico convertido, ou None se o valor
        de entrada for None, vazio, ou não puder ser convertido para
        número.
    """
    if valor is None:
        return None

    texto = str(valor).strip()

    if not texto:
        return None

    texto = (
        texto
        .replace("R$", "")
        .replace(" ", "")
    )

    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    else:
        partes = texto.split(".")
        if len(partes) > 2:
            texto = "".join(partes)

    try:
        return float(texto)
    except ValueError:
        return None


def ler_linhas_tratadas(caminho_csv, colunas_validas):
    """Gerador que percorre o CSV de remuneração linha a linha,
    descartando linhas totalmente vazias e mantendo, em cada linha,
    apenas os valores das colunas consideradas válidas (identificadas
    previamente por identificar_colunas_validas).

    Usar um gerador (yield) em vez de carregar tudo em memória permite
    processar arquivos muito grandes sem consumir memória excessiva.

    Parâmetros:
        caminho_csv (str): caminho do CSV de remuneração.
        colunas_validas (list[int]): índices das colunas (no CSV
            original) que devem ser mantidas em cada linha de saída.

    Produz (yield):
        list: valores da linha, já filtrados para conter apenas as
        colunas válidas, na mesma ordem de colunas_validas. Colunas
        vazias são representadas como string vazia "".
    """
    arquivo, leitor = preparar_csv(caminho_csv)

    try:
        next(leitor, None)

        for linha in leitor:
            if not any(
                not coluna_vazia(valor)
                for valor in linha
            ):
                continue

            linha_saida = []

            for indice in colunas_validas:
                valor = linha[indice] if indice < len(linha) else ""

                if coluna_vazia(valor):
                    linha_saida.append("")
                else:
                    linha_saida.append(valor)

            yield linha_saida

    finally:
        arquivo.close()


def identificar_colunas_monetarias(cabecalho):
    """Identifica quais colunas do cabeçalho representam valores
    monetários, com base em seus nomes conterem alguma das palavras-
    chave em PALAVRAS_MONETARIAS (ex: "remuneração", "gratificação",
    "irrf", "total" etc.), ignorando maiúsculas/minúsculas.

    Parâmetros:
        cabecalho (list[str]): lista com os nomes das colunas (já
            filtradas para as colunas válidas).

    Retorna:
        set[int]: conjunto com os índices (na lista `cabecalho`) das
        colunas identificadas como monetárias.
    """
    return {
        indice
        for indice, nome in enumerate(cabecalho)
        if nome is not None
        and any(
            palavra in str(nome).lower()
            for palavra in PALAVRAS_MONETARIAS
        )
    }


def calcular_largura_colunas(cabecalho):
    """Calcula a largura ideal (em unidades de coluna do Excel) para
    cada coluna da planilha final, com base em heurísticas sobre o
    nome da coluna (ex: colunas de nome de pessoa são mais largas,
    colunas de CPF têm largura fixa menor, colunas monetárias e de
    data têm larguras específicas), e um cálculo genérico baseado no
    tamanho do nome da coluna para os demais casos.

    Parâmetros:
        cabecalho (list[str]): lista com os nomes das colunas da
            planilha final.

    Retorna:
        list[int]: lista de larguras, uma por coluna, na mesma ordem do
        cabecalho.
    """
    larguras = []

    for nome in cabecalho:
        nome_coluna = (
            str(nome).lower().strip()
            if nome is not None
            else ""
        )

        if "nome" in nome_coluna:
            largura = 60
        elif "cpf" in nome_coluna:
            largura = 18
        elif "cargo" in nome_coluna:
            largura = 35
        elif "órgão" in nome_coluna or "orgao" in nome_coluna:
            largura = 30
        elif any(
            palavra in nome_coluna
            for palavra in PALAVRAS_MONETARIAS
        ):
            largura = 18
        elif "data" in nome_coluna:
            largura = 15
        else:
            largura = min(
                max(len(nome_coluna) + 4, 15),
                30
            )

        larguras.append(largura)

    return larguras


def criar_planilha_formatada(
    cabecalho,
    linhas,
    larguras,
    caminho_saida,
    logger
):
    """Gera o arquivo Excel (.xlsx) final, já formatado, a partir do
    cabeçalho, das linhas de dados (tipicamente vindas do gerador
    ler_linhas_tratadas) e das larguras de coluna calculadas.

    Usa o modo "write_only" do openpyxl (mais eficiente em memória para
    grandes volumes de dados, pois escreve direto em disco em vez de
    manter tudo em memória). Aplica formatação ao cabeçalho (negrito,
    texto branco, fundo azul, alinhamento centralizado, quebra de
    linha), congela a primeira linha (freeze_panes) e ajusta a largura
    de cada coluna. Para colunas identificadas como monetárias
    (identificar_colunas_monetarias), converte o valor para número
    (converter_valor_monetario) e aplica o formato de moeda
    (FORMATO_MOEDA); quando a conversão falha, mantém o valor original
    como texto. Registra no log o progresso a cada 50 mil registros
    processados e, ao final, define a área de autofiltro cobrindo todos
    os dados escritos.

    Parâmetros:
        cabecalho (list[str]): nomes das colunas da planilha.
        linhas (iterable): iterável (ou gerador) de linhas de dados, na
            mesma ordem/estrutura do cabeçalho.
        larguras (list[int]): largura de cada coluna, na mesma ordem do
            cabeçalho.
        caminho_saida (str): caminho onde o arquivo .xlsx final será
            salvo.
        logger (logging.Logger): logger usado para registrar o
            progresso e a conclusão da geração da planilha.
    """
    inicio = time.time()

    workbook = Workbook(write_only=True)
    planilha = workbook.create_sheet(title="Remuneração")

    fonte_cabecalho = Font(
        bold=True,
        color="FFFFFF"
    )

    preenchimento_cabecalho = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    alinhamento_cabecalho = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    borda_cabecalho = Border(
        bottom=Side(
            style="thin",
            color="FFFFFF"
        )
    )

    colunas_monetarias = identificar_colunas_monetarias(
        cabecalho
    )

    for indice, largura in enumerate(larguras, start=1):
        letra = get_column_letter(indice)
        planilha.column_dimensions[letra].width = largura

    planilha.freeze_panes = "A2"
    planilha.row_dimensions[1].height = 35

    linha_cabecalho = []

    for valor in cabecalho:
        celula = WriteOnlyCell(
            planilha,
            value=valor
        )
        celula.font = fonte_cabecalho
        celula.fill = preenchimento_cabecalho
        celula.alignment = alinhamento_cabecalho
        celula.border = borda_cabecalho
        linha_cabecalho.append(celula)

    planilha.append(linha_cabecalho)

    total = 0

    for linha in linhas:
        linha_saida = []

        for indice, valor in enumerate(linha):
            if indice in colunas_monetarias:
                numero = converter_valor_monetario(valor)

                if numero is not None:
                    celula = WriteOnlyCell(
                        planilha,
                        value=numero
                    )
                    celula.number_format = FORMATO_MOEDA
                    linha_saida.append(celula)
                else:
                    linha_saida.append(valor)
            else:
                linha_saida.append(valor)

        planilha.append(linha_saida)
        total += 1

        if total % 50000 == 0:
            logger.info(
                "%d registros processados.",
                total
            )

    ultima_letra = get_column_letter(
        len(larguras)
    )

    planilha.auto_filter.ref = (
        f"A1:{ultima_letra}{total + 1}"
    )

    workbook.save(caminho_saida)

    segundos = time.time() - inicio

    logger.info(
        "Planilha criada com %d registros em %.1f segundos.",
        total,
        segundos
    )


def extrair_zip(caminho_zip, pasta_extraida, logger):
    """Extrai todo o conteúdo do arquivo ZIP baixado para a pasta
    indicada.

    Parâmetros:
        caminho_zip (str): caminho do arquivo .zip a extrair.
        pasta_extraida (str): pasta de destino da extração.
        logger (logging.Logger): logger usado para registrar o início e
            a conclusão da extração.
    """
    logger.info("Extraindo o arquivo ZIP.")

    with zipfile.ZipFile(caminho_zip, "r") as zip_ref:
        zip_ref.extractall(pasta_extraida)

    logger.info("ZIP extraído com sucesso.")


def remover_arquivos_desnecessarios(pasta_extraida, logger):
    """Remove, da pasta onde o ZIP foi extraído, os arquivos que não são
    necessários para o processamento (ex: dados de afastamentos,
    cadastro e observações), mantendo apenas o que é relevante (o CSV
    de remuneração).

    Um arquivo é considerado desnecessário se seu nome (sem extensão,
    em minúsculas) contiver alguma das palavras em
    `nomes_desnecessarios`. Erros ao remover um arquivo específico (ex:
    permissão negada) são registrados como aviso no log, sem interromper
    o processo para os demais arquivos.

    Parâmetros:
        pasta_extraida (str): pasta onde o conteúdo do ZIP foi
            extraído.
        logger (logging.Logger): logger usado para registrar cada
            remoção e a contagem final.
    """
    nomes_desnecessarios = {
        "afastamentos",
        "cadastro",
        "observacoes",
        "observações",
    }

    removidos = 0

    for raiz, _, arquivos in os.walk(pasta_extraida):
        for arquivo in arquivos:
            nome_sem_extensao = os.path.splitext(arquivo)[0].lower()

            if any(
                nome in nome_sem_extensao
                for nome in nomes_desnecessarios
            ):
                caminho = os.path.join(raiz, arquivo)

                try:
                    os.remove(caminho)
                    removidos += 1
                    logger.info(
                        "Arquivo desnecessário removido: %s",
                        arquivo
                    )
                except OSError as erro:
                    logger.warning(
                        "Não foi possível remover %s: %s",
                        arquivo,
                        erro
                    )

    logger.info(
        "Limpeza concluída. %d arquivos removidos.",
        removidos
    )


def validar_periodo(ano, mes, tipo=None):
    """Valida se o período (ano/mês) solicitado é elegível para
    execução do pipeline, verificando: se o ano é conhecido
    (INDICES_ANOS), se o mês é um nome válido (MESES_OPCOES), se o
    período não está no futuro em relação à data atual, e se o mês está
    de fato disponível para aquele ano (meses_disponiveis) — relevante
    para o ano corrente, cujos meses futuros ainda não têm dados
    publicados.

    Parâmetros:
        ano (str): ano a validar.
        mes (str): mês a validar, por extenso.
        tipo: parâmetro mantido por compatibilidade de assinatura, não
            utilizado na validação.

    Levanta:
        ValueError: com uma mensagem explicando o motivo, caso o
            período seja inválido, esteja no futuro, ou não esteja
            disponível.
    """
    if ano not in INDICES_ANOS:
        raise ValueError("Ano inválido.")

    if mes not in MESES_OPCOES:
        raise ValueError("Mês inválido.")

    hoje = datetime.now()

    if (int(ano), MESES_OPCOES[mes]) > (hoje.year, hoje.month):
        raise ValueError(
            "Não é permitido executar para um mês futuro."
        )

    if mes not in meses_disponiveis(ano):
        raise ValueError(
            "O mês selecionado não está disponível para esse ano."
        )


def limpar_arquivos_extraidos(pasta):
    """Remove todo o conteúdo (arquivos e subpastas) de dentro da pasta
    de processamento (PASTA_EXTRAIDA), sem remover a própria pasta —
    usada para deixar a área de trabalho pronta para a próxima execução,
    sem acumular dados de execuções anteriores.

    Não faz nada se a pasta informada for None/vazia ou não existir.
    Erros ao remover itens individuais são silenciosamente ignorados
    (OSError), para não interromper a limpeza por causa de um único
    arquivo problemático.

    Parâmetros:
        pasta (str): caminho da pasta a ser esvaziada.
    """
    if not pasta or not os.path.isdir(pasta):
        return

    for nome in os.listdir(pasta):
        caminho = os.path.join(pasta, nome)
        try:
            if os.path.isfile(caminho) or os.path.islink(caminho):
                os.remove(caminho)
            elif os.path.isdir(caminho):
                shutil.rmtree(caminho, ignore_errors=True)
        except OSError:
            pass


def executar_pipeline_completo(
    ano,
    mes,
    tipo,
    log_queue,
):
    """Orquestra o pipeline completo do Robô SIAPE, do início ao fim:
    valida o período, cria a pasta e o logger desta execução, baixa o
    pacote ZIP (com retentativas), extrai seu conteúdo, remove arquivos
    desnecessários, localiza o CSV de remuneração, preserva uma cópia
    bruta desse CSV na pasta da execução, identifica as colunas válidas,
    calcula as larguras de coluna e, por fim, gera a planilha Excel
    tratada e formatada.

    É a função central chamada pela interface gráfica (através da
    thread Worker) para efetivamente executar toda a automação.

    Parâmetros:
        ano (str): ano da competência a processar.
        mes (str): mês da competência, por extenso.
        tipo (str): tipo do pacote de dados a baixar (ex:
            "Servidores_SIAPE").
        log_queue (queue.Queue): fila usada para repassar mensagens de
            log em tempo real para quem estiver consumindo (ex: a
            interface gráfica).

    Retorna:
        tuple: (caminho_saida, caminho_log) — caminho do arquivo Excel
        tratado gerado e caminho do arquivo de log desta execução.

    Levanta:
        ValueError: se o período informado for inválido (via
            validar_periodo).
        FileNotFoundError: se o pacote não estiver disponível para
            download, ou se o CSV de remuneração não for encontrado
            dentro do ZIP.
        RuntimeError: se o download falhar após todas as retentativas.
        Exception: qualquer outra exceção ocorrida durante o
            processamento é registrada no log (com stack trace) e
            relançada para o chamador tratar.

    Observações:
        Em qualquer cenário (sucesso ou falha), a pasta de trabalho
        temporária do download é removida, o conteúdo da pasta de
        processamento é limpo, e os handlers do logger desta execução
        são finalizados (flush + close), evitando vazamento de recursos
        entre execuções.
    """
    validar_periodo(ano, mes, tipo)

    os.makedirs(PASTA_SAIDA, exist_ok=True)

    pasta_execucao = criar_pasta_execucao(ano, mes)

    logger, caminho_log = configurar_logger(
        ano,
        mes,
        log_queue,
        pasta_execucao
    )

    pasta_download = None

    try:
        logger.info(
            "Início da execução para %s/%02d — pacote: %s.",
            ano,
            MESES_OPCOES[mes],
            tipo,
        )
        logger.info(
            "Pasta desta execução: %s",
            pasta_execucao,
        )

        caminho_zip, pasta_download = baixar_com_retentativas(
            ano,
            mes,
            tipo,
            logger,
        )

        os.makedirs(PASTA_EXTRAIDA, exist_ok=True)

        pasta_extraida = PASTA_EXTRAIDA

        extrair_zip(
            caminho_zip,
            pasta_extraida,
            logger
        )

        remover_arquivos_desnecessarios(
            pasta_extraida,
            logger
        )

        arquivo_remuneracao = encontrar_arquivo_remuneracao(
            pasta_extraida
        )

        logger.info(
            "CSV de remuneração encontrado: %s",
            arquivo_remuneracao
        )

        caminho_bruto = os.path.join(
            pasta_execucao,
            f"base_bruta_{ano}_{MESES_OPCOES[mes]:02d}.csv"
        )

        shutil.copy2(arquivo_remuneracao, caminho_bruto)
        logger.info(
            "Planilha bruta preservada em: %s",
            caminho_bruto
        )

        cabecalho, colunas_validas = identificar_colunas_validas(
            arquivo_remuneracao,
            logger
        )

        larguras = calcular_largura_colunas(
            cabecalho
        )

        caminho_saida = os.path.join(
            pasta_execucao,
            f"base_tratada_{ano}_{MESES_OPCOES[mes]:02d}.xlsx"
        )

        criar_planilha_formatada(
            cabecalho,
            ler_linhas_tratadas(
                arquivo_remuneracao,
                colunas_validas
            ),
            larguras,
            caminho_saida,
            logger
        )

        logger.info(
            "Execução concluída com sucesso."
        )

        return caminho_saida, caminho_log

    except Exception:
        logger.exception(
            "Erro durante a execução."
        )
        raise

    finally:
        if pasta_download:
            limpar_pasta_trabalho(pasta_download)

        limpar_arquivos_extraidos(PASTA_EXTRAIDA)

        for handler in list(logger.handlers):
            handler.flush()
            handler.close()
            logger.removeHandler(handler)